import asyncio
import base64
import io
import mimetypes
import os
import zipfile
from pathlib import Path
from typing import Optional
from xml.etree import ElementTree

from open_webui.env import (
    ENABLE_IMAGE_CONTENT_TYPE_EXTENSION_FALLBACK,
    LITE_OFFICE_FILE_CONTEXT_MAX_BYTES,
    LITE_OFFICE_FILE_CONTEXT_MAX_CHARS,
    LITE_OFFICE_SPREADSHEET_MAX_ROWS,
    LITE_OFFICE_SPREADSHEET_MAX_SHEETS,
    LITE_PDF_FILE_CONTEXT_MAX_BYTES,
    LITE_PDF_FILE_CONTEXT_MAX_CHARS,
    LITE_PDF_MAX_PAGES,
    LITE_TEXT_FILE_CONTEXT_MAX_BYTES,
)
from open_webui.models.files import Files
from open_webui.storage.provider import Storage


_TEXT_FILE_EXTENSIONS = {
    '.bat',
    '.c',
    '.cfg',
    '.conf',
    '.cpp',
    '.cs',
    '.css',
    '.csv',
    '.go',
    '.h',
    '.hpp',
    '.html',
    '.ini',
    '.java',
    '.js',
    '.json',
    '.jsonl',
    '.jsx',
    '.log',
    '.md',
    '.php',
    '.ps1',
    '.py',
    '.rb',
    '.rs',
    '.sh',
    '.sql',
    '.toml',
    '.ts',
    '.tsv',
    '.tsx',
    '.txt',
    '.xml',
    '.yaml',
    '.yml',
}

_IMAGE_MIME_FALLBACK = {
    '.webp': 'image/webp',
    '.png': 'image/png',
    '.jpg': 'image/jpeg',
    '.jpeg': 'image/jpeg',
    '.gif': 'image/gif',
    '.svg': 'image/svg+xml',
    '.bmp': 'image/bmp',
    '.tiff': 'image/tiff',
    '.tif': 'image/tiff',
    '.ico': 'image/x-icon',
    '.heic': 'image/heic',
    '.heif': 'image/heif',
    '.avif': 'image/avif',
}

_DOCX_CONTENT_TYPES = {
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
}

_XLSX_CONTENT_TYPES = {
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
}

_PPTX_CONTENT_TYPES = {
    'application/vnd.openxmlformats-officedocument.presentationml.presentation',
}

_PDF_CONTENT_TYPES = {
    'application/pdf',
    'application/x-pdf',
}

_MAX_OFFICE_XML_BYTES = 8 * 1024 * 1024
_SPREADSHEET_MAX_CELLS_PER_ROW = 50


def is_lite_text_file(filename: str, content_type: str | None) -> bool:
    content_type = content_type or ''
    if content_type.startswith('text/'):
        return True
    if content_type in {
        'application/json',
        'application/jsonl',
        'application/javascript',
        'application/x-javascript',
        'application/xml',
        'application/yaml',
        'application/x-yaml',
    }:
        return True
    return os.path.splitext(filename or '')[1].lower() in _TEXT_FILE_EXTENSIONS


def is_lite_office_file(filename: str, content_type: str | None) -> bool:
    extension = os.path.splitext(filename or '')[1].lower()
    content_type = content_type or ''
    return (
        extension in {'.docx', '.xlsx', '.pptx'}
        or content_type in _DOCX_CONTENT_TYPES
        or content_type in _XLSX_CONTENT_TYPES
        or content_type in _PPTX_CONTENT_TYPES
    )


def is_lite_pdf_file(filename: str, content_type: str | None) -> bool:
    extension = os.path.splitext(filename or '')[1].lower()
    content_type = content_type or ''
    return extension == '.pdf' or content_type in _PDF_CONTENT_TYPES


def is_lite_image_file(filename: str, content_type: str | None) -> bool:
    extension = os.path.splitext(filename or '')[1].lower()
    content_type = content_type or ''
    return content_type.startswith('image/') or extension in _IMAGE_MIME_FALLBACK


def get_lite_unsupported_file_message(filename: str, reason: str | None = None) -> tuple[int, str, str]:
    extension = os.path.splitext(filename or '')[1].lower()

    if reason == 'empty_upload':
        return 400, 'empty_upload', '不能上传空文件。请选择有内容的文件后再上传。'
    if reason == 'too_large':
        return 413, 'file_too_large', '文件超过 Render lite 的轻量处理限制。请缩小文件，或使用 Hugging Face full。'
    if reason == 'encrypted':
        return 422, 'encrypted_file', '这个文件似乎已加密，Render lite 无法读取。请解除密码后再上传。'
    if reason == 'empty_pdf':
        return (
            422,
            'pdf_no_text',
            '这个 PDF 没有检测到可复制文字，可能是扫描件。可以尝试转为图片后让视觉模型读取。',
        )
    if reason in {'empty', 'decode_failed', 'binary'}:
        return 422, 'no_readable_text', '这个文件没有提取到可读取文字。请确认文件内容不是图片、加密内容或损坏文件。'
    if reason in {'office_extract_failed', 'invalid_office_file'}:
        return 422, 'office_extract_failed', '这个 Office 文件无法读取。请重新另存为新版 Office 格式后再上传。'
    if reason == 'pdf_extract_failed':
        return 422, 'pdf_extract_failed', '这个 PDF 无法读取。请确认文件未损坏，或使用 Hugging Face full。'
    if reason == 'missing_dependency':
        return 500, 'missing_dependency', '当前服务缺少读取该文件所需的组件。请联系管理员处理。'

    if extension == '.doc':
        return 415, 'unsupported_legacy_office', '当前 Render lite 不支持 `.doc` 老版 Word 文件。请在本地另存为 `.docx` 后再上传。'
    if extension == '.xls':
        return 415, 'unsupported_legacy_office', '当前 Render lite 不支持 `.xls` 老版 Excel 文件。请在本地另存为 `.xlsx` 后再上传。'
    if extension == '.ppt':
        return 415, 'unsupported_legacy_office', '当前 Render lite 不支持 `.ppt` 老版 PowerPoint 文件。请在本地另存为 `.pptx` 后再上传。'
    if extension in {'.zip', '.rar', '.7z', '.tar', '.gz'}:
        return 415, 'unsupported_archive', '当前 Render lite 不支持直接上传压缩包。请先解压，再上传里面的支持格式文件。'
    if extension in {'.mp3', '.wav', '.m4a', '.flac', '.aac', '.ogg', '.mp4', '.mov', '.avi', '.mkv', '.webm'}:
        return 415, 'unsupported_media', '当前 Render lite 不支持音频或视频文件。请先在外部转写成文本后再上传。'
    if extension in {'.epub', '.odt', '.ods', '.rtf'}:
        return 415, 'unsupported_document', '当前 Render lite 暂不支持这种文档格式。请转换为 txt、md、docx、xlsx、pptx 或文本型 PDF 后再上传。'

    return 415, 'unsupported_file_type', '当前 Render lite 不支持上传这种文件格式。请转换为支持的格式后再上传。'


def validate_lite_file_upload(
    contents: bytes, filename: str, content_type: str | None
) -> tuple[bool, dict, dict | None]:
    if not contents:
        status_code, code, message = get_lite_unsupported_file_message(filename, 'empty_upload')
        return False, {}, {'status_code': status_code, 'code': code, 'message': message}

    if is_lite_image_file(filename, content_type):
        return True, {}, None

    if not (is_lite_text_file(filename, content_type) or is_lite_office_file(filename, content_type) or is_lite_pdf_file(filename, content_type)):
        status_code, code, message = get_lite_unsupported_file_message(filename)
        return False, {}, {'status_code': status_code, 'code': code, 'message': message}

    text_content, text_skip_reason = extract_lite_text_content(contents, filename, content_type)
    if text_content is not None:
        return (
            True,
            {
                'content': text_content,
                'content_type': 'text',
                'lite_text_context': True,
            },
            None,
        )

    reason = 'empty_pdf' if is_lite_pdf_file(filename, content_type) and text_skip_reason == 'empty' else text_skip_reason
    status_code, code, message = get_lite_unsupported_file_message(filename, reason)
    return False, {}, {'status_code': status_code, 'code': code, 'message': message}


def _append_limited_line(lines: list[str], line: str, state: dict, max_chars: int) -> bool:
    line = ' '.join(str(line).replace('\x00', '').split())
    if not line:
        return True

    separator_len = 1 if lines else 0
    remaining = max_chars - state['chars'] - separator_len
    if remaining <= 0:
        state['truncated'] = True
        return False

    if len(line) > remaining:
        lines.append(line[:remaining])
        state['chars'] = max_chars
        state['truncated'] = True
        return False

    lines.append(line)
    state['chars'] += len(line) + separator_len
    return True


def _xml_tag_name(tag: str) -> str:
    return tag.rsplit('}', 1)[-1]


def _docx_xml_lines(xml_bytes: bytes) -> list[str]:
    root = ElementTree.fromstring(xml_bytes)
    lines = []

    for paragraph in root.iter():
        if _xml_tag_name(paragraph.tag) != 'p':
            continue

        fragments = []
        for node in paragraph.iter():
            tag_name = _xml_tag_name(node.tag)
            if tag_name == 't' and node.text:
                fragments.append(node.text)
            elif tag_name == 'tab':
                fragments.append('\t')
            elif tag_name in {'br', 'cr'}:
                fragments.append('\n')

        line = ''.join(fragments).strip()
        if line:
            lines.extend(part.strip() for part in line.splitlines() if part.strip())

    return lines


def _extract_docx_text(contents: bytes) -> tuple[str | None, str | None]:
    lines: list[str] = []
    state = {'chars': 0, 'truncated': False}

    try:
        with zipfile.ZipFile(io.BytesIO(contents)) as archive:
            names = archive.namelist()
            xml_names = [
                name
                for name in names
                if name == 'word/document.xml'
                or name.startswith('word/header')
                or name.startswith('word/footer')
                or name in {'word/footnotes.xml', 'word/endnotes.xml'}
            ]
            xml_names.sort(key=lambda name: (0 if name == 'word/document.xml' else 1, name))

            if 'word/document.xml' not in xml_names:
                return None, 'invalid_office_file'

            for name in xml_names:
                info = archive.getinfo(name)
                if info.file_size > _MAX_OFFICE_XML_BYTES:
                    return None, 'too_large'

                for line in _docx_xml_lines(archive.read(name)):
                    if not _append_limited_line(lines, line, state, LITE_OFFICE_FILE_CONTEXT_MAX_CHARS):
                        break
                if state['truncated']:
                    break
    except Exception:
        return None, 'office_extract_failed'

    text = '\n'.join(lines).strip()
    if not text:
        return None, 'empty'
    if state['truncated']:
        text = f'{text}\n\n[Content truncated for Render lite context limit.]'
    return text, None


def _extract_xlsx_text(contents: bytes) -> tuple[str | None, str | None]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return None, 'missing_dependency'

    lines: list[str] = []
    state = {'chars': 0, 'truncated': False}

    try:
        workbook = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        try:
            for sheet_index, sheet in enumerate(workbook.worksheets):
                if sheet_index >= LITE_OFFICE_SPREADSHEET_MAX_SHEETS:
                    state['truncated'] = True
                    break

                if not _append_limited_line(lines, f'[Sheet: {sheet.title}]', state, LITE_OFFICE_FILE_CONTEXT_MAX_CHARS):
                    break

                for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if row_index > LITE_OFFICE_SPREADSHEET_MAX_ROWS:
                        if not _append_limited_line(
                            lines,
                            f'[Sheet truncated after {LITE_OFFICE_SPREADSHEET_MAX_ROWS} rows.]',
                            state,
                            LITE_OFFICE_FILE_CONTEXT_MAX_CHARS,
                        ):
                            break
                        state['truncated'] = True
                        break

                    values = []
                    for value in row:
                        if value is None:
                            continue
                        text = str(value).strip()
                        if text:
                            values.append(text)
                        if len(values) >= _SPREADSHEET_MAX_CELLS_PER_ROW:
                            values.append('[row truncated]')
                            break

                    if values and not _append_limited_line(
                        lines, '\t'.join(values), state, LITE_OFFICE_FILE_CONTEXT_MAX_CHARS
                    ):
                        break

                if state['truncated']:
                    break
        finally:
            workbook.close()
    except Exception:
        return None, 'office_extract_failed'

    text = '\n'.join(lines).strip()
    if not text:
        return None, 'empty'
    if state['truncated']:
        text = f'{text}\n\n[Content truncated for Render lite context limit.]'
    return text, None


def _office_part_number(name: str, prefix: str, suffix: str) -> int:
    stem = name.removeprefix(prefix).removesuffix(suffix)
    return int(stem) if stem.isdigit() else 0


def _extract_pptx_text(contents: bytes) -> tuple[str | None, str | None]:
    lines: list[str] = []
    state = {'chars': 0, 'truncated': False}

    try:
        with zipfile.ZipFile(io.BytesIO(contents)) as archive:
            slide_names = [
                name
                for name in archive.namelist()
                if name.startswith('ppt/slides/slide') and name.endswith('.xml')
            ]
            slide_names.sort(key=lambda name: _office_part_number(name, 'ppt/slides/slide', '.xml'))

            if not slide_names:
                return None, 'invalid_office_file'

            for index, name in enumerate(slide_names, start=1):
                info = archive.getinfo(name)
                if info.file_size > _MAX_OFFICE_XML_BYTES:
                    return None, 'too_large'

                slide_lines = _docx_xml_lines(archive.read(name))
                if slide_lines and not _append_limited_line(
                    lines, f'[Slide {index}]', state, LITE_OFFICE_FILE_CONTEXT_MAX_CHARS
                ):
                    break

                for line in slide_lines:
                    if not _append_limited_line(lines, line, state, LITE_OFFICE_FILE_CONTEXT_MAX_CHARS):
                        break
                if state['truncated']:
                    break
    except Exception:
        return None, 'office_extract_failed'

    text = '\n'.join(lines).strip()
    if not text:
        return None, 'empty'
    if state['truncated']:
        text = f'{text}\n\n[Content truncated for Render lite context limit.]'
    return text, None


def extract_lite_office_content(
    contents: bytes, filename: str, content_type: str | None
) -> tuple[str | None, str | None]:
    if not is_lite_office_file(filename, content_type):
        return None, None
    if len(contents) > LITE_OFFICE_FILE_CONTEXT_MAX_BYTES:
        return None, 'too_large'

    extension = os.path.splitext(filename or '')[1].lower()
    content_type = content_type or ''
    if extension == '.docx' or content_type in _DOCX_CONTENT_TYPES:
        return _extract_docx_text(contents)
    if extension == '.xlsx' or content_type in _XLSX_CONTENT_TYPES:
        return _extract_xlsx_text(contents)
    if extension == '.pptx' or content_type in _PPTX_CONTENT_TYPES:
        return _extract_pptx_text(contents)
    return None, None


def extract_lite_pdf_content(contents: bytes, filename: str, content_type: str | None) -> tuple[str | None, str | None]:
    if not is_lite_pdf_file(filename, content_type):
        return None, None
    if len(contents) > LITE_PDF_FILE_CONTEXT_MAX_BYTES:
        return None, 'too_large'

    try:
        from pypdf import PdfReader
    except Exception:
        return None, 'missing_dependency'

    lines: list[str] = []
    state = {'chars': 0, 'truncated': False}

    try:
        reader = PdfReader(io.BytesIO(contents), strict=False)
        if reader.is_encrypted:
            try:
                decrypt_result = reader.decrypt('')
            except Exception:
                decrypt_result = 0
            if not decrypt_result:
                return None, 'encrypted'

        page_count = len(reader.pages)
        for page_index, page in enumerate(reader.pages[:LITE_PDF_MAX_PAGES], start=1):
            page_text = page.extract_text() or ''
            page_lines = [line.strip() for line in page_text.splitlines() if line.strip()]
            if not page_lines:
                continue

            if not _append_limited_line(lines, f'[Page {page_index}]', state, LITE_PDF_FILE_CONTEXT_MAX_CHARS):
                break
            for line in page_lines:
                if not _append_limited_line(lines, line, state, LITE_PDF_FILE_CONTEXT_MAX_CHARS):
                    break
            if state['truncated']:
                break

        if page_count > LITE_PDF_MAX_PAGES:
            state['truncated'] = True
    except Exception:
        return None, 'pdf_extract_failed'

    text = '\n'.join(lines).strip()
    if not text:
        return None, 'empty'
    if state['truncated']:
        text = f'{text}\n\n[Content truncated for Render lite context limit.]'
    return text, None


def extract_lite_text_content(contents: bytes, filename: str, content_type: str | None) -> tuple[str | None, str | None]:
    if not is_lite_text_file(filename, content_type):
        text_content, text_skip_reason = extract_lite_office_content(contents, filename, content_type)
        if text_content is not None or text_skip_reason is not None:
            return text_content, text_skip_reason
        return extract_lite_pdf_content(contents, filename, content_type)

    if len(contents) > LITE_TEXT_FILE_CONTEXT_MAX_BYTES:
        return None, 'too_large'
    if b'\x00' in contents:
        return None, 'binary'

    for encoding in ('utf-8', 'utf-8-sig', 'latin-1'):
        try:
            return contents.decode(encoding), None
        except UnicodeDecodeError:
            continue
    return None, 'decode_failed'


async def get_lite_file_sources(items: list[dict] | None, user=None) -> list[dict]:
    if not items or user is None:
        return []

    sources = []
    seen = set()
    for item in items:
        if not isinstance(item, dict):
            continue
        file_id = item.get('id') or item.get('url')
        if not file_id or file_id in seen:
            continue
        seen.add(file_id)

        file = await Files.get_file_by_id(file_id)
        if not file or (file.user_id != user.id and user.role != 'admin'):
            continue

        data = file.data or {}
        content = data.get('content')
        if not isinstance(content, str) or not content:
            continue

        meta = file.meta or {}
        source_name = meta.get('name') or item.get('name') or file.filename or file.id
        sources.append(
            {
                'source': {'id': file.id, 'name': source_name, 'type': 'file'},
                'document': [content],
                'metadata': [
                    {
                        'file_id': file.id,
                        'name': source_name,
                        'source': source_name,
                        'content_type': meta.get('content_type'),
                    }
                ],
            }
        )

    return sources


async def get_image_base64_from_file_id(id: str, user=None) -> Optional[str]:
    file = await Files.get_file_by_id(id)
    if not file or user is None:
        return None

    if file.user_id != user.id and user.role != 'admin':
        return None

    try:
        meta = file.meta or {}
        content_type = meta.get('content_type') if isinstance(meta, dict) else None
        if content_type and not content_type.startswith('image/'):
            return None

        file_path = await asyncio.to_thread(Storage.get_file, file.path)
        file_path = Path(file_path)
        if not file_path.is_file():
            return None

        content_type = content_type or mimetypes.guess_type(file_path.name)[0]
        if not content_type and ENABLE_IMAGE_CONTENT_TYPE_EXTENSION_FALLBACK:
            content_type = _IMAGE_MIME_FALLBACK.get(file_path.suffix.lower())
        if not content_type or not content_type.startswith('image/'):
            return None

        with open(file_path, 'rb') as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
        return f'data:{content_type};base64,{encoded_string}'
    except Exception:
        return None


async def get_image_base64_from_url(url: str, user=None) -> Optional[str]:
    if url.startswith('data:image/'):
        return url
    if url.startswith('http'):
        return None
    return await get_image_base64_from_file_id(url, user=user)
